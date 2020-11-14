from openpyxl import Workbook, load_workbook
from random import shuffle
import sys
import argparse

def arrange_seat(input_file, number_of_rooms):
    workbook = load_workbook(filename=input_file, data_only=True)
    try:
        students = workbook['students']
    except KeyError:
        print("student worksheet does not present in the file.")
    try:
        stats = workbook['stats']
    except KeyError:
        print("stats worksheet does not present in the file.")
    number_of_students = int(stats['A2'].value)
    index_list = list(range(1, number_of_students + 1))
    shuffle(index_list)
    counter = 0
    try:
        for i in range(2, int(number_of_rooms) + 2):
            ws = workbook[stats[f"C{i}"].value]
            for col in range(ord(stats[f"D{i}"].value), ord(stats[f"E{i}"].value) + 1):
                for row in range(5, int(stats[f"F{i}"].value)):
                    if counter > int(number_of_students):
                        break
                    if ws[f"{chr(col)}{row}"].value == 'x':
                        ws[f"{chr(col)}{row}"] = students[f"A{index_list[counter]}"].value
                        counter += 1
    except IndexError:
        pass
    workbook.save(filename=f"arranged_{input_file}")
    workbook.close()

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('-i', '--input_file')
    parser.add_argument('-n', '--rooms')
    args = parser.parse_args()

    arrange_seat(args.input_file, args.rooms)

if __name__ == '__main__':
    try:
        main()
    except:
        print("Something wrong with provided input file")
        sys.exit(0)